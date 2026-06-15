#!/usr/bin/env python3
"""Generate the Excel run report from the real Phase 8 + Phase 9 data (Phase 10).

Writes reports/referral-safety-report.xlsx with five sheets — Summary (KPIs +
charts), Referrals (per-referral detail), Human review, Exceptions, and the Audit
trail — driven entirely by report_model.build_model() (no hand-typed figures).

    python build_report.py [--out reports/referral-safety-report.xlsx]

Requires openpyxl. ALL DATA IS SYNTHETIC. OpenMRS is a mock EPR/EMR.
"""

from __future__ import annotations

import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import report_model  # noqa: E402

_REPO_ROOT = report_model._REPO_ROOT
DEFAULT_OUT = os.path.join(_REPO_ROOT, "reports", "referral-safety-report.xlsx")

NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT = "D9E1F2"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"
GREY = "808080"

HEAD = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, color=NAVY, size=16)
SUB = Font(italic=True, color=GREY, size=9)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(cell):
    cell.font = HEAD
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = CENTER
    cell.border = BORDER


def _table(ws, top_row, headers, rows, col_widths=None, status_col=None):
    """Write a bordered table starting at top_row; returns the next free row."""
    for c, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=top_row, column=c, value=h))
    r = top_row + 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if status_col and c == status_col:
                cell.fill = PatternFill("solid", fgColor=_status_fill(str(val)))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r + 1


def _status_fill(label: str) -> str:
    if "Created" in label:
        return GREEN
    if "Rejected" in label:
        return AMBER
    if "exception" in label.lower():
        return RED
    return "FFFFFF"


def _banner(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws["A2"] = subtitle
    ws["A2"].font = SUB
    ws["A3"] = "Synthetic data only - OpenMRS is a mock EPR/EMR - not a live NHS system."
    ws["A3"].font = SUB


def build_summary(wb, m):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    _banner(ws, "Clinical Referral Safety Automation - Run Report",
            "Generated from the " + m["generated_from"] + ".")

    k = m["kpis"]
    cards = [
        ("Referrals processed", k["total_referrals"]),
        ("Created in OpenMRS", k["created_in_openmrs"]),
        ("  - fully automated", k["auto_created"]),
        ("  - human-approved", k["human_approved_created"]),
        ("Routed to human review", k["routed_to_human_review"]),
        ("Rejected at review", k["rejected_no_record"]),
        ("Exceptions (BE / SE)", k["exceptions"]),
        ("Auto-created WITH a safety flag", k["auto_created_with_safety_flag"]),
        ("Audit rows written", k["audit_rows"]),
    ]
    row = 5
    ws.cell(row=row, column=1, value="Key metric").font = HEAD
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=row, column=2, value="Value").font = HEAD
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=row, column=2).alignment = CENTER
    for label, value in cards:
        row += 1
        ws.cell(row=row, column=1, value=label).border = BORDER
        vc = ws.cell(row=row, column=2, value=value)
        vc.border = BORDER
        vc.alignment = CENTER
        vc.font = BOLD
        if label.startswith("Auto-created WITH"):
            vc.fill = PatternFill("solid", fgColor=GREEN if value == 0 else RED)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12

    # Outcome breakdown + pie chart.
    orow = 5
    ws.cell(row=orow, column=4, value="Final outcome").font = HEAD
    ws.cell(row=orow, column=4).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=orow, column=5, value="Count").font = HEAD
    ws.cell(row=orow, column=5).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=orow, column=5).alignment = CENTER
    items = list(m["final_status_counts"].items())
    for i, (label, count) in enumerate(items, start=1):
        ws.cell(row=orow + i, column=4, value=label).border = BORDER
        c = ws.cell(row=orow + i, column=5, value=count)
        c.border = BORDER
        c.alignment = CENTER
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10

    pie = PieChart()
    pie.title = "Final outcomes (15 referrals)"
    data = Reference(ws, min_col=5, min_row=orow, max_row=orow + len(items))
    cats = Reference(ws, min_col=4, min_row=orow + 1, max_row=orow + len(items))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.height, pie.width = 7.5, 12
    ws.add_chart(pie, "D" + str(orow + len(items) + 2))

    # Safety-catch table + bar chart, lower down.
    srow = 17
    ws.cell(row=srow, column=1, value="Safety hazard caught and routed to a human").font = HEAD
    ws.cell(row=srow, column=1).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=srow, column=2, value="Referrals").font = HEAD
    ws.cell(row=srow, column=2).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=srow, column=2).alignment = CENTER
    sitems = list(m["safety_catches"].items())
    for i, (label, count) in enumerate(sitems, start=1):
        ws.cell(row=srow + i, column=1, value=label).border = BORDER
        c = ws.cell(row=srow + i, column=2, value=count)
        c.border = BORDER
        c.alignment = CENTER

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Safety hazards caught (by family)"
    bar.legend = None
    bdata = Reference(ws, min_col=2, min_row=srow, max_row=srow + len(sitems))
    bcats = Reference(ws, min_col=1, min_row=srow + 1, max_row=srow + len(sitems))
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.height, bar.width = 8, 12
    ws.add_chart(bar, "D" + str(srow))


def build_referrals(wb, m):
    ws = wb.create_sheet("Referrals")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Per-referral detail (15)", "Each referral, its decision path, and terminal outcome.")
    headers = ["Referral", "Patient (synthetic)", "Extraction", "Conf.", "Match result",
               "Bot decision", "Reviewer decision", "Final outcome", "OpenMRS encounter"]
    rows = []
    for r in m["referrals"]:
        rows.append([
            r["referral_id"], r["patient_name"], r["extraction_method"], r["confidence"],
            r["match_result"], r["bot_decision"].replace("_", " ").title(),
            r["reviewer_decision"] or "-", r["final_label"],
            (r["encounter_uuid"][:8] + "..." if r["encounter_uuid"] else "-"),
        ])
    _table(ws, 5, headers, rows,
           col_widths=[10, 20, 12, 8, 20, 18, 16, 28, 16], status_col=8)


def build_reviews(wb, m):
    ws = wb.create_sheet("Human review")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Human-in-the-loop review (10)",
            "Clinician decisions that changed the final outcome (Phase 9).")
    headers = ["Referral", "Patient", "Why routed (reason codes)", "Reviewer",
               "Decision", "Outcome", "Rationale"]
    rows = [[r["referral_id"], r["patient_name"], " | ".join(r["reason_codes"]),
             r["reviewer"], r["reviewer_decision"], r["final_label"], r["rationale"]]
            for r in m["human_reviews"]]
    _table(ws, 5, headers, rows, col_widths=[10, 16, 30, 16, 12, 26, 50], status_col=6)


def build_exceptions(wb, m):
    ws = wb.create_sheet("Exceptions")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Exceptions (2)", "Business vs System exception separation (REFramework).")
    headers = ["Referral", "Source", "Reason codes", "Bot decision", "Final outcome"]
    rows = [[r["referral_id"], r["source_file"], " | ".join(r["reason_codes"]),
             r["bot_decision"].replace("_", " ").title(), r["final_label"]]
            for r in m["exceptions"]]
    _table(ws, 5, headers, rows, col_widths=[10, 26, 22, 20, 24], status_col=5)


def build_audit(wb, m):
    ws = wb.create_sheet("Audit trail")
    ws.sheet_view.showGridLines = False
    _banner(ws, "Append-only audit trail (" + str(len(m["audit_rows"])) + " rows)",
            "Every system action: who / what / before-after / when / why.")
    headers = ["Timestamp (UTC)", "Referral", "Action", "Status", "Match result", "Detail"]
    rows = [[a["timestamp"], a["referral_id"], a["action"], a["status"],
             a["match_result"], a["detail"]] for a in m["audit_rows"]]
    _table(ws, 5, headers, rows, col_widths=[26, 10, 26, 28, 18, 60])


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate the Excel run report (Phase 10).")
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    m = report_model.build_model()
    wb = Workbook()
    build_summary(wb, m)
    build_referrals(wb, m)
    build_reviews(wb, m)
    build_exceptions(wb, m)
    build_audit(wb, m)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    k = m["kpis"]
    print(f"Wrote {args.out}")
    print(f"  {k['total_referrals']} referrals | {k['created_in_openmrs']} created "
          f"({k['auto_created']} auto + {k['human_approved_created']} human-approved) | "
          f"{k['rejected_no_record']} rejected | {k['exceptions']} exceptions | "
          f"{k['audit_rows']} audit rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
