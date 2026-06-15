#!/usr/bin/env python3
"""Phase 10 acceptance gate — reporting reflects the real run.

Asserts that the report model reconciles with the raw run artifacts (the audit
log and the Phase 3 oracles), that the core safety invariant holds, and that both
deliverables (Excel + HTML dashboard) are produced and well-formed.

    python validate_acceptance.py

ALL DATA IS SYNTHETIC.
"""

from __future__ import annotations

import json
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_dashboard  # noqa: E402
import build_report  # noqa: E402
import report_model  # noqa: E402

_REPO_ROOT = report_model._REPO_ROOT
ORACLE_DIR = os.path.join(_REPO_ROOT, "data", "expected-outcomes")

PASS, FAIL = "PASS", "FAIL"
results: list[tuple[str, str, str]] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    results.append((PASS if ok else FAIL, name, detail))


def main() -> int:
    m = report_model.build_model()
    k = m["kpis"]

    check("model covers all 15 referrals", k["total_referrals"] == 15,
          f"total={k['total_referrals']}")

    # Outcomes reconcile with the headline run: 9 created (3 auto + 6 human), 4 rejected, 2 exceptions.
    check("outcomes: 9 created (3 auto + 6 human-approved), 4 rejected, 2 exceptions",
          k["created_in_openmrs"] == 9 and k["auto_created"] == 3
          and k["human_approved_created"] == 6 and k["rejected_no_record"] == 4
          and k["exceptions"] == 2,
          json.dumps({x: k[x] for x in ("created_in_openmrs", "auto_created",
                     "human_approved_created", "rejected_no_record", "exceptions")}))

    # Core safety invariant: nothing carrying a safety flag was auto-created.
    check("safety invariant: 0 referrals auto-created with a safety flag",
          k["auto_created_with_safety_flag"] == 0,
          f"={k['auto_created_with_safety_flag']}")

    # The 'created' count derived from the model must equal the count derived
    # independently from the append-only audit log.
    audit = m["audit_rows"]
    audit_created = sum(1 for a in audit if a["status"] == "REFERRAL_CREATED_IN_OPENMRS")
    check("reconciles with audit log: created count matches the audit trail",
          audit_created == k["created_in_openmrs"],
          f"audit={audit_created} model={k['created_in_openmrs']}")
    check("every referral appears in the audit trail",
          {a["referral_id"] for a in audit} >= {r["referral_id"] for r in m["referrals"]})

    # Cross-check the automated + exception outcomes against the Phase 3 oracles
    # (reporting must reflect the *correct* run, not just any run).
    oracle_ok = True
    for r in m["referrals"]:
        op = os.path.join(ORACLE_DIR, f"{r['referral_id']}.expected.json")
        if not os.path.exists(op):
            continue
        exp = json.load(open(op, encoding="utf-8-sig"))
        if r["bot_decision"] != exp["expected_bot_decision"]:
            oracle_ok = False
    check("bot decisions match the Phase 3 oracles", oracle_ok)

    # Both artifacts are produced and well-formed.
    tmp = tempfile.mkdtemp(prefix="phase10-accept-")
    xlsx = os.path.join(tmp, "report.xlsx")
    htmlf = os.path.join(tmp, "dashboard.html")
    sys.argv = ["build_report.py", "--out", xlsx]
    build_report.main()
    sys.argv = ["build_dashboard.py", "--out", htmlf]
    build_dashboard.main()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx)
        sheets_ok = wb.sheetnames == ["Summary", "Referrals", "Human review",
                                      "Exceptions", "Audit trail"]
    except Exception as exc:  # noqa: BLE001
        sheets_ok = False
        check("excel: opens with the 5 expected sheets", False, str(exc))
    else:
        check("excel: opens with the 5 expected sheets", sheets_ok, str(wb.sheetnames))

    doc = open(htmlf, encoding="utf-8").read()
    check("dashboard: self-contained HTML with inline SVG charts and the headline figures",
          "<svg" in doc and "Created in OpenMRS" in doc and "Synthetic data only" in doc
          and "{{" not in doc, f"size={len(doc)}")

    npass = sum(1 for r in results if r[0] == PASS)
    print("\nPhase 10 - Reporting + dashboard: acceptance gate\n" + "=" * 55)
    for status, name, detail in results:
        print(f"[{status}] {name}" + (f"  ({detail})" if detail and status == FAIL else ""))
    print("=" * 55)
    print(f"{npass}/{len(results)} checks passed")
    return 0 if npass == len(results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
